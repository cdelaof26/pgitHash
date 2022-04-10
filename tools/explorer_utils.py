from tools.lang import SELECT_THE_DIRECTORY
from tools.lang import DO_YOU_WANT_EXPLORE_SUBDIRECTORIES
from tools.lang import DO_YOU_WANT_SPLIT_THE_DB_BY_DIRECTORY
from tools.lang import ENTER_A_NAME_FOR_YOUR_DB
from tools.lang import SELECT_THE_OLDER_BD
from tools.lang import SELECT_THE_NEWER_BD
from tools.lang import ENTER_A_NAME_FOR_YOUR_CDB
from tools.lang import SELECT_THE_CDB
from tools.lang import EXPLORING_DIR
from tools.lang import CANNOT_READ_CONTENTS
from tools.lang import SUCCEED
from tools.lang import CANNOT_WRITE_FILE
from tools.lang import FILE_CREATED_AT
from tools.lang import NOT_FILES_FOUND
from tools.lang import XLSX_DOESNT_MEET_COLUMNS
from tools.lang import CREATING
from tools.lang import M_OBJECTS_OF_N_OBJECTS
from tools.lang import BAD_DB_ORDER
from tools.lang import PROCESSED
from tools.lang import M_FILES_OF_N_FILES
from tools.lang import FILE_NOT_FOUND

from tools.FileHash import FileHash
from tools.FileHash import FILE_MOVED
from tools.FileHash import FILE_CREATED
from tools.FileHash import FILE_DELETED
from tools.FileHash import FILE_RENAMED
from tools.FileHash import FILE_MODIFIED
from tools.FileHash import FILE_RENAMED_AND_MOVED
from tools.FileHash import FILE_MODIFIED_AND_MOVED

from tools.util import choose
from tools.util import print_status

from tools.file_utils import file_browser
from tools.file_utils import create_directories
from tools.file_utils import copy_file
from tools.file_utils import create_file_hash

from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from zipfile import BadZipFile

# Utilities related to exploration of filesystem and creation of pseudo databases

CWD = Path.cwd()


# User_input: Path, bool, str -> "/home", True, "db"
# Output:     list            -> [["/home"], PosixPath("/home/cdelaof/db.xlsx"), True]
#
def setup_pdb_creation() -> list:
    print(SELECT_THE_DIRECTORY)
    directory = file_browser(allow_files=False)

    print(DO_YOU_WANT_EXPLORE_SUBDIRECTORIES)
    explore_subdirectories = choose(["1", "2"], [True, False])

    divide_by_directory = False

    if explore_subdirectories:
        print(DO_YOU_WANT_SPLIT_THE_DB_BY_DIRECTORY)
        divide_by_directory = choose(["1", "2"], [True, False])

    pdb_name = input(ENTER_A_NAME_FOR_YOUR_DB) + ".xlsx"
    pdb_path = Path.cwd().joinpath(pdb_name)

    return [[directory], pdb_path, explore_subdirectories, divide_by_directory]


# User_input: Path, Path, str -> "C:\\db.xlsx", "C:\\db1.xlsx", "cdb"
# Output:     list            -> [WindowsPath("C:\\db.xlsx"), WindowsPath("C:\\db1.xlsx"), WindowsPath("C:\\cdb.xlsx")]
#
def setup_pdb_comparison() -> list:
    print(SELECT_THE_OLDER_BD)
    older_pdb = file_browser(allow_directories=False, allowed_extensions=[".xlsx"])

    print(SELECT_THE_NEWER_BD)
    newer_pdb = file_browser(allow_directories=False, allowed_extensions=[".xlsx"])

    comparison_pdb_name = input(ENTER_A_NAME_FOR_YOUR_CDB) + ".xlsx"
    comparison_pdb_path = Path.cwd().joinpath(comparison_pdb_name)

    return [older_pdb, newer_pdb, comparison_pdb_path]


# User_input: Path, Path, str -> "C:\\db.xlsx", "C:\\db1.xlsx", "cdb"
# Output:     list            -> [WindowsPath("C:\\db.xlsx"), WindowsPath("C:\\db1.xlsx"), WindowsPath("C:\\cdb.xlsx")]
#
def setup_ach_comparison() -> list:
    print(SELECT_THE_OLDER_BD)
    older_pdb = file_browser(allow_directories=False, allowed_extensions=[".xlsx"])

    print(SELECT_THE_NEWER_BD)
    newer_pdb = file_browser(allow_directories=False, allowed_extensions=[".xlsx"])

    print(SELECT_THE_CDB)
    comparison_pdb_path = file_browser(allow_directories=False, allowed_extensions=[".xlsx"])

    return [older_pdb, newer_pdb, comparison_pdb_path]


# cp
def setup_cp() -> list:
    print(SELECT_THE_DIRECTORY)
    origin = file_browser(allow_files=False)

    print(SELECT_THE_DIRECTORY)
    destiny = file_browser(allow_files=False)

    return [origin, destiny]


# Input: list, list, bool
#
# Notes: Uses reference
#
def explore_dir(directories, files, explore_subdirectories, hash_func, blacklist_extensions):
    # I think it's a lot of processing just to display it nicely, maybe I will change it.
    print_status(EXPLORING_DIR, directories[0])

    try:
        for item in directories[0].iterdir():
            if item.name in blacklist_extensions or item.suffix in blacklist_extensions:
                continue

            # In order to prevent an infinite process, cwd is excluded to be scanned
            # likely to happen with divide_by_directory option enabled
            if item == CWD:
                continue

            if explore_subdirectories and item.is_dir():
                directories.append(item)
            elif item.is_file():
                file_hash = None
                if hash_func is not None:
                    file_hash = create_file_hash(item, hash_func)
                
                files.append(FileHash(str(item), file_hash))
    except (PermissionError, FileNotFoundError) as e:
        print(CANNOT_READ_CONTENTS % str(directories[0]))
        print(e)
        print()

    directories.pop(0)


# Input:  list     -> ["Ingredient", "Price"], [FileHash(), FileHash(), FileHash()], Path("[...]")
#
def write_xlsx(columns, data, file_path):
    if not data:
        print(NOT_FILES_FOUND)
    else:
        workbook = Workbook()
        sheet = workbook.active

        for column, column_name in enumerate(columns):
            sheet.cell(row=1, column=(column + 1), value=column_name)

        for i, file in enumerate(data):
            # i + 2 since first row is for column labels
            sheet.cell(row=(i + 2), column=1, value=str(file.get_file_hash()))
            sheet.cell(row=(i + 2), column=2, value=str(file.get_file_path()))
            sheet.cell(row=(i + 2), column=3, value=str(file.get_file_stat()))

        while True:
            try:
                # Just in case if file_path has " on its name
                file_path = str(file_path).replace("\"", "\\\"")

                workbook.save(file_path)

                print(SUCCEED + FILE_CREATED_AT.format(str(file_path)))

                workbook.close()

                break
            except FileNotFoundError as e:
                print(CANNOT_WRITE_FILE % (file_path, e))
                if choose(["1", "2"], [False, True]):
                    break


# Too way complicated explain it xd
#
def retrieve_workbook_origin_path(sheet, file_path=None) -> list:
    # if there is not sheet, then is created
    workbook = None

    if file_path:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except BadZipFile:
            print(CANNOT_READ_CONTENTS % str(file_path))
            return list()

    # Origin path is only available for p-databases
    # cp-databases uses column 3 for file_stat

    # Basically does this:
    #       "Origin_path: Path_name".split(": ")
    #       ['Origin_path', 'Path_name']
    # If path origin_path_split[1] is found, then is returned
    #
    origin_path_split = sheet.cell(row=1, column=3).value
    origin_path = None
    matched_columns = True

    if origin_path_split is not None:
        origin_path_split = origin_path_split.split(": ")

        if len(origin_path_split) == 2:
            origin_path = origin_path_split[1]
        else:
            # If is needed data_origin_path but not found, then file
            # Doesn't meet the columns needed
            matched_columns = False

    if file_path:
        workbook.close()

    return [origin_path, matched_columns]


# Input:  list, Path -> ["Ingredient", "Price"], "/root/c.xlsx"
# Output: list       -> [<tools.FileHash.FileHash object at [...]>,
#                        <tools.FileHash.FileHash object at [...]>,
#                        ...]
#
def retrieve_workbook_objects(columns, file_path, retrieve_origin_path) -> list:
    try:
        workbook = load_workbook(file_path)
    except BadZipFile:
        print(CANNOT_READ_CONTENTS % str(file_path))
        return list()

    sheet = workbook.active
    matched_columns = True

    files = list()
    origin_path = None

    for column, column_name in enumerate(columns):
        matched_columns = matched_columns and (sheet.cell(row=1, column=(column + 1)).value == column_name)

    if retrieve_origin_path:
        origin_path, matched_columns = retrieve_workbook_origin_path(sheet)

    if not matched_columns:
        print(XLSX_DOESNT_MEET_COLUMNS % (str(file_path), columns, origin_path is not None))
        workbook.close()
        if retrieve_origin_path:
            return [files, origin_path]
        return files

    total_objects = sheet.max_row - 1
    created_objects = 1

    # Starts in 2 because we don't want the column name as object
    # (The other +1 in 2 is because openpyxl cells start with 1 :v)
    for row in range(2, sheet.max_row + 1):
        print_status(CREATING, M_OBJECTS_OF_N_OBJECTS.format(created_objects, total_objects))

        file_hash = str(sheet.cell(row=row, column=1).value)
        file_path = Path(sheet.cell(row=row, column=2).value)
        file_stat = str(sheet.cell(row=row, column=3).value)

        files.append(FileHash(file_path, file_hash, file_stat))

        created_objects += 1

    print()

    if retrieve_origin_path:
        return [files, origin_path]

    return files


# Input:  list, list
# Output: list
#
# Notes: example in README
#
def compare_data(older_db_data, newer_db_data, older_db_origin_path, newer_db_origin_path) -> list:
    cdb_data = list()

    files_processed = 1
    total_files = len(older_db_data)

    while older_db_data:
        print_status(PROCESSED, M_FILES_OF_N_FILES.format(files_processed, total_files))

        old_data = older_db_data[0]
        remove_old_data_from_list = False

        # This returns a part of the original path
        #   Let's say:
        #           older_db_origin_path = "/Users/cdelaof26/Desktop"
        #           old_data.get_file_path() = "/Users/cdelaof26/Desktop/pgitHash/settings/File.txt"
        #
        # To compare relative paths to catch file movement, we just compare after origin and excludes file's name
        #           old_data_partial_path = "/pgitHash/settings"
        #           new_data_partial_path = "/pgitHash/"
        #
        #        -> same_path = False
        #
        # The full path cannot be used since mount point for newer and older databases is not the same.
        #
        old_data_path = old_data.get_file_path()
        old_data_partial_path = str(old_data_path).replace(older_db_origin_path, "")
        old_data_partial_path = old_data_partial_path.replace(old_data_path.name, "")

        for i_new_data, new_data in enumerate(newer_db_data):
            same_names = Path(old_data.get_file_path()).stem == Path(new_data.get_file_path()).stem

            # Code explanation at line 285
            new_data_path = new_data.get_file_path()
            new_data_partial_path = str(new_data_path).replace(newer_db_origin_path, "")
            new_data_partial_path = new_data_partial_path.replace(new_data_path.name, "")

            same_path = old_data_partial_path == new_data_partial_path

            # If hashes are the same
            if old_data.get_file_hash() == new_data.get_file_hash():

                if same_names and not same_path:
                    # If hashes and names are the same but paths not, is considered as moved
                    old_data.set_file_stat(FILE_MOVED)
                    cdb_data.append(old_data)
                if not same_names and same_path:
                    # If hashes and paths are the same but the names are not, is considered as renamed
                    old_data.set_file_stat(FILE_RENAMED)
                    cdb_data.append(old_data)
                if not same_names and not same_path:
                    # If hashes are the same but names and paths are not, is considered as RenamedAndMoved
                    old_data.set_file_stat(FILE_RENAMED_AND_MOVED)
                    cdb_data.append(old_data)

                # If hashes, paths and names are the same is considered as unmodified
                # so, not added to list of differences

                newer_db_data.pop(i_new_data)
                remove_old_data_from_list = True
                break
            elif same_names:
                if same_path:
                    # If hashes are not the same, but paths and names are, then is considered as modified
                    old_data.set_file_stat(FILE_MODIFIED)
                else:
                    # If hashes and paths are not the same, but names are, then is considered as ModifiedAndMoved
                    old_data.set_file_stat(FILE_MODIFIED_AND_MOVED)

                cdb_data.append(old_data)
                newer_db_data.pop(i_new_data)
                remove_old_data_from_list = True
                break

        if not remove_old_data_from_list:
            # If a file was not found in newer data is considered as deleted
            old_data.set_file_stat(FILE_DELETED)
            cdb_data.append(old_data)

        # Removes old files
        older_db_data.pop(0)
        files_processed += 1

    # Any non-deleted file on newer data is considered as created
    while newer_db_data:
        new_data = newer_db_data[0]

        # Code explanation at line 285
        # To push changes is needed change the path to the older one
        new_data_path = new_data.get_file_path()
        # To join it, it can't start with / or \, that's why [1:]
        new_data_partial_path = str(new_data_path).replace(newer_db_origin_path, "")[1:]
        # Here name is not remove since we need it

        new_data_path = Path(older_db_origin_path).joinpath(new_data_partial_path)

        new_data.set_file_stat(FILE_CREATED)
        new_data.set_file_path(new_data_path)

        cdb_data.append(newer_db_data[0])

        newer_db_data.pop(0)

    return cdb_data


# Input: str, list
#
def search_file_by_hash(file_hash, data) -> FileHash:
    for dat in data:
        if dat.get_file_hash() == file_hash:
            return dat

    print(FILE_NOT_FOUND % file_hash)
    return FileHash(Path.cwd(), "", "")


# Input: str, list
#
def search_file_by_name(file_name, data) -> FileHash:
    for dat in data:
        if dat.get_file_path().name == file_name:
            return dat

    print(FILE_NOT_FOUND % file_name)
    return FileHash(Path.cwd(), "", "")


# Input: Path, str, list, Path, Path
#
def move_file(old_file_path, old_file_hash, newer_db_data, newer_db_origin_path, older_db_origin_path,
              search_by_hash=True):
    if search_by_hash:
        new_file = search_file_by_hash(old_file_hash, newer_db_data)
    else:
        new_file = search_file_by_name(old_file_hash, newer_db_data)

    # Code explanation at line 285
    #
    # Those lines are needed to create the new partial path of the file
    # Here we are just getting the directories between newer_db_origin_path and file's name
    #
    new_data_path = new_file.get_file_path()
    # To join it, it can't start with / or \, that's why [1:]
    new_partial_path = str(new_data_path).replace(newer_db_origin_path, "")[1:]
    new_partial_path = new_partial_path.replace(new_data_path.name, "")

    new_path = Path(older_db_origin_path).joinpath(new_partial_path)

    create_directories(new_path)

    copy_file(old_file_path, new_path, move_file=True)

    if new_file.get_file_hash() != "":
        newer_db_data.remove(new_file)


# Input: Path, list
#
def delete_and_copy_file(old_file_path, newer_db_data, delete_data=True):
    if old_file_path.exists():
        old_file_path.unlink()

    if not old_file_path.exists():
        new_file = search_file_by_name(old_file_path.name, newer_db_data)
        copy_file(new_file.get_file_path(), old_file_path.parent)
        if delete_data and new_file.get_file_hash() != "":
            newer_db_data.remove(new_file)


# Input: list, list, Path, Path
# Too way complicated, sorry
#
def apply_changes(cdb_data, newer_db_data, newer_db_origin_path, older_db_origin_path):
    total_files = len(cdb_data)
    files_processed = 1

    # Little verification
    # Since there's the possibility of using older_db as newer_db_data
    # cdb_data would cause weird errors and cause data lost, so there's a
    # validation to prevent that.
    #
    # That is done by checking if an element from cdb_data contains older_db_origin_path
    if str(older_db_origin_path) not in str(cdb_data[0].get_file_path()):
        print(BAD_DB_ORDER)
        return

    for file in cdb_data:
        print_status(PROCESSED, M_FILES_OF_N_FILES.format(files_processed, total_files))

        old_file_hash = file.get_file_hash()
        old_file_path = file.get_file_path()
        old_file_stat = file.get_file_stat()

        # FILE_MOVED                DONE
        # FILE_CREATED              DONE
        # FILE_DELETED              DONE
        # FILE_RENAMED              DONE
        # FILE_MODIFIED             DONE
        # FILE_RENAMED_AND_MOVED    DONE
        # FILE_MODIFIED_AND_MOVED   DONE

        if old_file_stat == FILE_MOVED:
            move_file(old_file_path, old_file_hash, newer_db_data, newer_db_origin_path, older_db_origin_path)

        if old_file_stat == FILE_CREATED:
            new_file = search_file_by_hash(old_file_hash, newer_db_data)
            # Before copying it, is verified if parent exist
            create_directories(old_file_path.parent)

            copy_file(new_file.get_file_path(), old_file_path.parent)
            if new_file.get_file_hash() != "":
                newer_db_data.remove(new_file)

        if old_file_stat == FILE_DELETED:
            if old_file_path.exists():
                old_file_path.unlink()

        if old_file_stat == FILE_RENAMED:
            if old_file_path.exists():
                new_file = search_file_by_hash(old_file_hash, newer_db_data)
                new_file_name = new_file.get_file_path().name
                old_file_path.replace(old_file_path.with_name(new_file_name))
                if new_file.get_file_hash() != "":
                    newer_db_data.remove(new_file)

        if old_file_stat == FILE_MODIFIED:
            delete_and_copy_file(old_file_path, newer_db_data)

        if old_file_stat == FILE_RENAMED_AND_MOVED:
            if old_file_path.exists():
                new_file = search_file_by_hash(old_file_hash, newer_db_data)
                new_file_name = new_file.get_file_path().name
                old_file_path = old_file_path.replace(old_file_path.with_name(new_file_name))

            if old_file_path.exists():
                move_file(old_file_path, old_file_hash, newer_db_data, newer_db_origin_path, older_db_origin_path)

        if old_file_stat == FILE_MODIFIED_AND_MOVED:
            delete_and_copy_file(old_file_path, newer_db_data, delete_data=False)

            if old_file_path.exists():
                move_file(old_file_path, old_file_path.name, newer_db_data, newer_db_origin_path, older_db_origin_path,
                          search_by_hash=False)

        files_processed += 1
