from tools.lang import WELCOME
from tools.lang import MENU
from tools.lang import WHAT_DO_YOU_WANT_TO_DO
from tools.lang import SELECT_THE_DIRECTORY
from tools.lang import DO_YOU_WANT_EXPLORE_SUBDIRECTORIES
from tools.lang import SUCCEED

from tools.util import choose
from tools.util import file_browser

from tools.explorer_utils import explore_dir

from pathlib import Path

print(WELCOME)

while True:
    print(MENU)
    print(WHAT_DO_YOU_WANT_TO_DO)
    operation = choose(["1", "2", "3", "E"])

    if operation == "1":
        print(SELECT_THE_DIRECTORY)
        directory = file_browser(allow_files=False)

        print(DO_YOU_WANT_EXPLORE_SUBDIRECTORIES)
        explore_subdirectories = choose(["1", "2"], [True, False])

        file_name = input("Enter a name for your database\n> ") + ".xlsx"
        file_path = Path.cwd().joinpath(file_name)

        directories = [directory]
        files = list()

        while directories:
            explore_dir(directories, files, explore_subdirectories)
        print()

        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        columns_matches = 0

        sheet.cell(row=1, column=1, value="Hash")
        sheet.cell(row=1, column=2, value="File_path")

        for i, file in enumerate(files):
            # i + 2 since first row is for column labels
            sheet.cell(row=(i + 2), column=1, value=file.get_hash())
            sheet.cell(row=(i + 2), column=2, value=file.get_file_path())

        workbook.save(file_path)

        print(SUCCEED % str(Path.cwd()))
    elif operation == "E":
        break

print("Bye")
